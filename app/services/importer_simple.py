"""
Importador simplificado para archivos CSV/Excel de Mercado Pago
Solo extrae columnas esenciales y clasifica por REAL_AMOUNT
"""
import csv
import hashlib
import uuid
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import load_workbook

from app.models.enums import TransactionType, SourceType, TransactionStatus, Currency
from app.services.classifier import TransactionClassifier
from app.services.categorizer import TransactionCategorizer
from app.services.merchant_normalizer import MerchantNormalizer
from app.repositories import get_db


class MercadoPagoImporterSimple:
    """
    Importador simplificado de Mercado Pago
    Extrae solo columnas esenciales y clasifica por REAL_AMOUNT
    Con deduplicación persistente por SOURCE_ID
    """

    # Columnas conocidas del archivo de Mercado Pago
    # Nota: Se mantendrán TODAS las columnas originales del CSV
    KNOWN_COLUMNS = {
        'SOURCE_ID': 'external_id',
        'TRANSACTION_DATE': 'operation_date',
        'REAL_AMOUNT': 'real_amount',
        'TRANSACTION_CURRENCY': 'currency',
        'TRANSACTION_TYPE': 'transaction_type_raw',
        'DESCRIPTION': 'description',
        'STORE_NAME': 'store_name',
        'POS_NAME': 'pos_name',
        'PAYMENT_METHOD': 'payment_method',
        'PAYMENT_METHOD_TYPE': 'payment_method_type',
        'INSTALLMENTS': 'installments',
    }

    def __init__(self):
        self.classifier = TransactionClassifier()
        self.categorizer = TransactionCategorizer()
        self.merchant_normalizer = MerchantNormalizer()
        self.db = get_db()

        # Set para trackear SOURCE_IDs del archivo actual (no persistente)
        self.current_file_source_ids = set()

    def import_file(
        self,
        file_path: Path,
        source_type: SourceType,
    ) -> Dict[str, Any]:
        """
        Importa archivo de Mercado Pago y retorna resultado con transacciones

        Returns:
            Dict con summary y lista de transacciones
        """
        # Limpiar set de SOURCE_IDs para este archivo
        self.current_file_source_ids.clear()

        # Parse file
        rows = self._parse_file(file_path)

        # Process cada fila
        transactions = []
        processed = 0
        failed = 0
        duplicates = 0
        review_required = 0

        for idx, row in enumerate(rows):
            try:
                # Verificar duplicado por SOURCE_ID en el archivo actual
                source_id = row.get('SOURCE_ID')
                if source_id and source_id in self.current_file_source_ids:
                    duplicates += 1
                    continue

                # Parsear fila a transacción
                tx_dict = self._parse_row(row, source_type)

                if tx_dict:
                    # Marcar SOURCE_ID como procesado en este archivo
                    if source_id:
                        self.current_file_source_ids.add(source_id)

                    transactions.append(tx_dict)
                    processed += 1

                    # Verificar si necesita revisión (baja confianza)
                    if tx_dict.get('category_confidence', 1.0) < 0.6:
                        review_required += 1

            except Exception as e:
                failed += 1
                print(f"Error procesando fila {idx}: {e}")

        # Generar ID único para este batch
        batch_id = str(uuid.uuid4())

        # Guardar transacciones en memoria asociadas al batch
        if 'import_batches' not in self.db.data:
            self.db.data['import_batches'] = {}

        self.db.data['import_batches'][batch_id] = {
            'id': batch_id,
            'total_rows': len(rows),
            'processed_rows': processed,
            'duplicated_rows': duplicates,
            'failed_rows': failed,
            'metadata': {
                'review_required': review_required
            },
            'transactions': transactions,
            'imported_at': datetime.now().isoformat()
        }
        self.db.save()

        # Construir respuesta con estructura batch
        return {
            "success": True,
            "batch": {
                "id": batch_id,
                "total_rows": len(rows),
                "processed_rows": processed,
                "duplicated_rows": duplicates,
                "failed_rows": failed,
                "metadata": {
                    "review_required": review_required
                }
            }
        }

    def _parse_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Parse CSV o Excel y retorna lista de diccionarios"""
        if file_path.suffix.lower() in ['.xlsx', '.xls']:
            return self._parse_excel(file_path)
        else:
            return self._parse_csv(file_path)

    def _parse_excel(self, file_path: Path) -> List[Dict[str, Any]]:
        """Parse Excel usando openpyxl"""
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active

        # Leer headers de primera fila
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter)

        # Normalizar nombres de columna (uppercase, strip)
        headers = [
            str(h).strip().upper() if h is not None else f'COL_{i}'
            for i, h in enumerate(headers)
        ]

        # Leer filas de datos
        data_rows = []
        for row_values in rows_iter:
            row_dict = {}
            for header, value in zip(headers, row_values):
                # Convertir None y strings vacíos a None
                if value is None or value == '':
                    row_dict[header] = None
                else:
                    row_dict[header] = value
            data_rows.append(row_dict)

        wb.close()
        return data_rows

    def _parse_csv(self, file_path: Path) -> List[Dict[str, Any]]:
        """Parse CSV con múltiples intentos de encoding"""
        # Probar diferentes encodings
        for encoding in ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']:
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as f:
                    # Detectar delimitador
                    sample = f.read(8192)
                    f.seek(0)

                    # Intentar detectar delimitador
                    sniffer = csv.Sniffer()
                    try:
                        dialect = sniffer.sniff(sample, delimiters=',;\t')
                        delimiter = dialect.delimiter
                    except csv.Error:
                        delimiter = ','

                    reader = csv.DictReader(f, delimiter=delimiter)

                    # Normalizar nombres de columna
                    if reader.fieldnames:
                        reader.fieldnames = [
                            str(h).strip().upper() for h in reader.fieldnames
                        ]

                    # Leer todas las filas
                    data_rows = []
                    for row in reader:
                        # Convertir strings vacíos a None
                        normalized_row = {
                            k: (None if v == '' else v) for k, v in row.items()
                        }
                        data_rows.append(normalized_row)

                    return data_rows

            except (UnicodeDecodeError, Exception):
                continue

        raise ValueError("No se pudo decodificar el archivo CSV")

    def _parse_row(
        self, row: Dict[str, Any], source_type: SourceType
    ) -> Optional[Dict[str, Any]]:
        """
        Parse una fila a formato de transacción simplificado

        Returns:
            Dict con formato de respuesta API o None si falla
        """
        # Extraer campos requeridos
        source_id = row.get('SOURCE_ID')
        real_amount = self._parse_float(row.get('REAL_AMOUNT'))
        description = str(row.get('DESCRIPTION') or 'Sin descripción').strip()
        operation_date = self._parse_date(row.get('TRANSACTION_DATE'))
        payment_method = row.get('PAYMENT_METHOD', '').lower()
        payment_method_type = row.get('PAYMENT_METHOD_TYPE', '').lower()

        # Validar campos requeridos
        if real_amount is None or operation_date is None:
            return None

        # EXCEPCIÓN: Si PAYMENT_METHOD = credit_card y REAL_AMOUNT > 0, ignorar
        # (evita que las transferencias/reembolsos con tarjeta aparezcan como ingresos falsos)
        if payment_method == 'credit_card' and real_amount > 0:
            return None

        # Clasificar tipo por REAL_AMOUNT
        # Positivo = ingreso, Negativo = egreso/gasto
        if real_amount > 0:
            tx_type = TransactionType.INCOME
        elif real_amount < 0:
            tx_type = TransactionType.EXPENSE
        else:
            # Monto = 0, ignorar
            return None

        # Valor absoluto del monto
        abs_amount = abs(real_amount)

        # Extraer merchant
        merchant_raw = self._extract_merchant(row, description)
        merchant_normalized = (
            self.merchant_normalizer.normalize(merchant_raw)
            if merchant_raw
            else None
        )
        merchant_display = merchant_normalized or merchant_raw or description

        # Categorizar
        category, category_confidence, category_reason = self.categorizer.categorize(
            merchant_normalized=merchant_normalized,
            description=description,
            payment_method=row.get('PAYMENT_METHOD'),
            transaction_type_str=tx_type.value,
        )

        # Determinar status - NO asignar "pendiente" automáticamente
        status = TransactionStatus.CONFIRMED

        # Moneda - SIEMPRE en ARS (Pesos)
        currency_code = row.get('TRANSACTION_CURRENCY') or 'ARS'
        currency_str = 'ARS'  # Forzar ARS

        # Método de pago
        payment_method = row.get('PAYMENT_METHOD') or ''

        # Cuotas
        installments = self._parse_int(row.get('INSTALLMENTS'))

        # Mantener TODAS las columnas originales del CSV
        all_original_columns = {}
        for col_name, col_value in row.items():
            if col_value is not None and col_value != '':
                all_original_columns[col_name] = col_value

        # Construir dict de transacción para respuesta
        return {
            "operation_date": operation_date.isoformat() if operation_date else None,
            "description": description,
            "merchant": merchant_display,
            "merchant_normalized": merchant_normalized,
            "amount": abs_amount,  # Valor absoluto
            "real_amount": real_amount,  # Valor original con signo
            "currency": currency_str,  # Siempre ARS
            "transaction_type": tx_type.value,  # "ingreso" o "gasto"
            "suggested_category_id": category or "sin_categoria",
            "status": status.value,  # Siempre "confirmada"
            "payment_method": payment_method,
            "payment_method_type": payment_method_type,
            "installments": installments,
            "source_id": source_id,
            "category_confidence": category_confidence,
            "all_columns": all_original_columns,  # TODAS las columnas originales
        }

    def _extract_merchant(
        self, row: Dict[str, Any], description: str
    ) -> Optional[str]:
        """Extrae merchant con orden de prioridad"""
        # Prioridad 1: STORE_NAME
        if row.get('STORE_NAME'):
            return str(row['STORE_NAME']).strip()

        # Prioridad 2: POS_NAME
        if row.get('POS_NAME'):
            return str(row['POS_NAME']).strip()

        # Prioridad 3: Extraer de DESCRIPTION
        if description:
            return self.merchant_normalizer.extract_from_description(description)

        return None

    def _parse_float(self, value: Any) -> Optional[float]:
        """Parse valor como float"""
        if value is None:
            return None
        try:
            # Manejar representaciones de string
            if isinstance(value, str):
                # Remover símbolos de moneda y espacios
                value = value.strip().replace('$', '').replace(',', '').replace(' ', '')
            return float(value)
        except (ValueError, TypeError):
            return None

    def _parse_int(self, value: Any) -> Optional[int]:
        """Parse valor como int"""
        if value is None:
            return None
        try:
            if isinstance(value, str):
                value = value.strip()
            return int(float(value))  # float() primero para manejar "1.0"
        except (ValueError, TypeError):
            return None

    def _parse_date(self, date_value: Any) -> Optional[datetime]:
        """Parse valor de fecha a datetime"""
        if date_value is None:
            return None

        try:
            if isinstance(date_value, datetime):
                return date_value

            # Intentar parsear strings de fecha
            if isinstance(date_value, str):
                # Formatos comunes de fecha
                for fmt in [
                    '%Y-%m-%d %H:%M:%S',
                    '%Y-%m-%d',
                    '%d/%m/%Y %H:%M:%S',
                    '%d/%m/%Y',
                    '%Y/%m/%d %H:%M:%S',
                    '%Y/%m/%d',
                    '%d-%m-%Y %H:%M:%S',
                    '%d-%m-%Y',
                ]:
                    try:
                        return datetime.strptime(date_value.strip(), fmt)
                    except ValueError:
                        continue

            # Intentar formato ISO como último recurso
            return datetime.fromisoformat(str(date_value))
        except:
            return None
